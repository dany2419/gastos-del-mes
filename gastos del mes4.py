import customtkinter as ctk
from tkinter import ttk, messagebox
from datetime import datetime
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import os, sys            # ⬅️  ya estaba os, ahora añadimos sys

# --- Carpeta del ejecutable (o del .py si estás en desarrollo) ---
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

DB_PATH = os.path.join(BASE_DIR, "gastos.db")

# Configuración inicial de customtkinter
ctk.set_appearance_mode("dark")  # Puedes cambiar a "light" o "system"
ctk.set_default_color_theme("blue")  # Temas: "blue", "green", "dark-blue"


# ------------------- INTERFAZ -------------------
ventana = ctk.CTk()
ventana.title("Registro de Gastos")
ventana.geometry("900x700")
ventana.resizable(False, False)

frame_entrada = ctk.CTkFrame(ventana)
frame_entrada.pack(pady=10, padx=200, fill="x")

# Campos de entrada
ctk.CTkLabel(frame_entrada, text="Monto ($):", font=("Helvetica", 12, "bold"))\
    .grid(row=0, column=0, padx=5, pady=5)
entrada_monto = ctk.CTkEntry(frame_entrada)
entrada_monto.grid(row=0, column=1, padx=5, pady=5)

ctk.CTkLabel(frame_entrada, text="Categoría:", font=("Helvetica", 12))\
    .grid(row=1, column=0, padx=5, pady=5)
combo_categoria = ctk.CTkComboBox(frame_entrada, values=[
    "Telefonía/Cable", "Servicios", "Tarjetas/Préstamos", "Otros"
])
combo_categoria.grid(row=1, column=1, padx=5, pady=5)
combo_categoria.set("Telefonía/Cable")

ctk.CTkLabel(frame_entrada, text="Descripción:", font=("Helvetica", 12))\
    .grid(row=2, column=0, padx=5, pady=5)
entrada_descripcion = ctk.CTkEntry(frame_entrada, width=300)
entrada_descripcion.grid(row=2, column=1, padx=5, pady=5)

ctk.CTkLabel(frame_entrada, text="Mes:", font=("Helvetica", 12))\
    .grid(row=3, column=0, padx=5, pady=5)
combo_mes = ctk.CTkComboBox(frame_entrada, values=[
    "Enero","Febrero","Marzo","Abril","Mayo","Junio",
    "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"
])
combo_mes.grid(row=3, column=1, padx=5, pady=5)
combo_mes.set("Abril")

# ------------------- BASE DE DATOS -------------------



def crear_base():
    conexion = sqlite3.connect("DB_PATH")
    cursor = conexion.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS gastos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha TEXT,
            monto REAL,
            categoria TEXT,
            descripcion TEXT,
            mes TEXT
        )
    """)
    conexion.commit()
    conexion.close()

def actualizar_tabla():
    conexion = sqlite3.connect("DB_PATH")
    cursor = conexion.cursor()
    try:
        cursor.execute("ALTER TABLE gastos ADD COLUMN mes TEXT")
        conexion.commit()
    except sqlite3.OperationalError:
        pass
    conexion.close()

# Inicializar base    
crear_base()
actualizar_tabla()

# ------------------- FUNCIONES -------------------
def mostrar_gastos():
    try:
        conexion = sqlite3.connect("DB_PATH")
        cursor   = conexion.cursor()
        cursor.execute(
            "SELECT id, fecha, monto, categoria, descripcion, mes FROM gastos"
        )
        registros = cursor.fetchall()
        conexion.close()
    except Exception as e:
        messagebox.showerror("Error al leer la base", str(e))
        return  # ⚠️ Salir para que no se ejecute el 'for' sin registros

    # --- Si todo salió bien, actualizamos la tabla ---
    tabla.delete(*tabla.get_children())
    for fila in registros:
        tabla.insert("", "end", values=fila)

    # Limpiar selección vieja (opcional)
    if hasattr(tabla, "id_sel"):
        delattr(tabla, "id_sel")


def agregar_gasto():
    raw = entrada_monto.get().replace(",", ".")
    try:
        monto = float(raw)
        fecha = datetime.now().strftime("%d/%m/%Y")
        conexion = sqlite3.connect("DB_PATH")
        cursor = conexion.cursor()
        cursor.execute(
            "INSERT INTO gastos (fecha, monto, categoria, descripcion, mes) VALUES (?,?,?,?,?)",
            (fecha, monto, combo_categoria.get(), entrada_descripcion.get(), combo_mes.get())
        )
        conexion.commit()
        conexion.close()
        entrada_monto.delete(0, "end")
        entrada_descripcion.delete(0, "end")
        mostrar_gastos()
    except ValueError:
        messagebox.showerror("Error", "Ingresa un número válido (ej. 1234.56).")

def seleccionar_gasto(event):
    sel = tabla.focus()
    vals = tabla.item(sel, "values")
    if vals:
        entrada_monto.delete(0, "end")
        entrada_monto.insert(0, vals[2])
        combo_categoria.set(vals[3])
        entrada_descripcion.delete(0, "end")
        entrada_descripcion.insert(0, vals[4])
        combo_mes.set("")  # opcional, no se muestra
        tabla.id_sel = vals[0]

def modificar_gasto():
    # ⬇️ 1) Chequeo rápido
    if not hasattr(tabla, "id_sel"):
        messagebox.showwarning("Atención", "Seleccioná un gasto primero.")
        return

    # ⬇️ 2) Lógica original
    try:
        monto = float(entrada_monto.get().replace(",", "."))
        conexion = sqlite3.connect("DB_PATH")
        cursor = conexion.cursor()
        cursor.execute(
            "UPDATE gastos SET monto=?, categoria=?, descripcion=?, mes=? WHERE id=?",
            (monto, combo_categoria.get(), entrada_descripcion.get(),
             combo_mes.get(), tabla.id_sel)
        )
        conexion.commit()
        conexion.close()
        mostrar_gastos()
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo modificar: {e}")

def eliminar_gasto():
    # ⬇️ 1) Chequeo rápido
    if not hasattr(tabla, "id_sel"):
        messagebox.showwarning("Atención", "Seleccioná un gasto primero.")
        return

    # ⬇️ 2) Lógica original
    try:
        conexion = sqlite3.connect("DB_PATH")
        cursor = conexion.cursor()
        cursor.execute("DELETE FROM gastos WHERE id=?", (tabla.id_sel,))
        conexion.commit()
        conexion.close()
        mostrar_gastos()
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo eliminar: {e}")



def exportar_a_excel():
    conexion = sqlite3.connect("DB_PATH")
    cursor = conexion.cursor()
    cursor.execute("SELECT fecha, monto, categoria, descripcion, mes FROM gastos")
    datos = cursor.fetchall()
    conexion.close()
    libro = Workbook()
    hoja = libro.active; hoja.title = "Gastos"
    hoja.merge_cells("A1:E1")
    hoja["A1"] = f"Resumen de {combo_mes.get()}"
    hoja["A1"].font = Font(size=14, bold=True)
    hoja["A1"].alignment = Alignment(horizontal="center")
    hoja["A1"].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    encabezados = ["Fecha","Monto","Categoría","Descripción","Mes"]
    hoja.append(encabezados)
    for i, txt in enumerate(encabezados, 1):
        cel = hoja.cell(row=2, column=i)
        cel.font = Font(bold=True)
        cel.fill = PatternFill(fill_type="solid", fgColor="B4C6E7")
        cel.alignment = Alignment(horizontal="center")
    for fila in datos:
        hoja.append(fila)
    for col in hoja.columns:
        lc = get_column_letter(col[0].column)
        m = max((len(str(c.value)) for c in col if c.value), default=0)
        hoja.column_dimensions[lc].width = m + 4
    nombre = f"gastos_{combo_mes.get().lower()}.xlsx"
    libro.save(nombre)
    messagebox.showinfo("Exportado", f"Gastos guardados en {nombre}")

# ------------------- TABLA -------------------
# Nota: La tabla sigue siendo de ttk.Treeview porque customtkinter no tiene su propia versión
tabla = ttk.Treeview(ventana,
    columns=("ID","Fecha","Monto","Categoría","Descripción"), show="headings")
for col, txt in [("ID","ID"),("Fecha","Fecha"),("Monto","Monto"),
                 ("Categoría","Categoría"),("Descripción","Descripción")]:
    tabla.heading(col, text=txt)
    tabla.column(col, width=80 if col=="ID" else 120)
tabla.pack(pady=10, padx=10, fill="both", expand=True)
tabla.bind("<ButtonRelease-1>", seleccionar_gasto)

# Estilo para la tabla (mejorado para customtkinter)
style = ttk.Style()
style.theme_use("default")
style.configure("Treeview",
    background="#2a2d2e",
    foreground="white",
    rowheight=25,
    fieldbackground="#2a2d2e",
    bordercolor="#343638",
    borderwidth=0)
style.map('Treeview', background=[('selected', '#22559b')])
style.configure("Treeview.Heading",
    background="#565b5e",
    foreground="white",
    relief="flat")
style.map("Treeview.Heading",
    background=[('active', '#3484F0')])

# ------------------- BOTONES -------------------
frame_botones = ctk.CTkFrame(ventana)
frame_botones.pack(pady=10, padx=30, fill="x")

ctk.CTkButton(frame_botones, text="Agregar Gasto", command=agregar_gasto,
              fg_color="green", hover_color="#2d7d46", width=200, height=40)\
    .grid(row=0, column=0, padx=5, pady=5)
ctk.CTkButton(frame_botones, text="Modificar Gasto", command=modificar_gasto,
              fg_color="#ce9ede", hover_color="#dc99f2", width=200, height=40)\
    .grid(row=0, column=1, padx=5, pady=5)
ctk.CTkButton(frame_botones, text="Eliminar Gasto", command=eliminar_gasto,
              fg_color="#DB4437", hover_color="#ba3a30", width=200, height=40)\
    .grid(row=0, column=2, padx=5, pady=5)
ctk.CTkButton(frame_botones, text="Exportar a Excel", command=exportar_a_excel,
              fg_color="#4c66af", hover_color="#3a5496", width=200, height=40)\
    .grid(row=0, column=3, padx=5, pady=5)

# Iniciar mostrando datos
mostrar_gastos()
ventana.mainloop()